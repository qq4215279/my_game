#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
扫描 Datas 目录下的 Luban 数据 Excel，增量更新 __tables__.xlsx 表定义。

合并策略（以 input 为唯一键）：
  - input 未变化：保留数据维护者在 __tables__.xlsx 中手动修改的其他字段
  - 扫描到新增表：追加一行，使用默认命名规则填充
  - 扫描中已不存在的表：删除对应行

命名规则（仅新增行使用）：
  - full_name  : {模块.}{PascalCase子表名}Table
  - value_type : {模块.}{PascalCase子表名}
  - output     : {snake_case子表名}
  - 模块（Java 子包名）优先级：
      1. 文件名含 '.'：去掉 '#' 后第一个 '.' 前为包名（如 #demo.demo.xlsx -> item）
      2. 否则取 Datas 子目录名（如 item/foo.xlsx -> item）
  - 生成 Java 包：com.mumu.game.luban.config.{模块}.*
  - 子表名 xxx_base / base_xxx 归一化为 xxx
  - 单 sheet 且 sheet 名为默认名（Sheet1 等）时，用 Excel 文件名作为子表名（去掉 #，取最后一个 '.' 后段）
  - 文件重命名但逻辑子表不变（如 #Item2.xlsx -> #demo.demo.xlsx）时，保留原行并迁移命名空间
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[ERROR] 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# 元数据文件，不参与扫描
SKIP_FILES = {"__tables__.xlsx", "__beans__.xlsx", "__enums__.xlsx"}
# 跳过目录：表定义已在 Defines/*.xml 中维护（如 Luban 官方 test 示例）
SKIP_DIRS = {"test"}
# Luban 数据文件后缀
DATA_SUFFIXES = {".xlsx", ".xls", ".csv"}
# Excel 默认 sheet 名（大小写不敏感）
DEFAULT_SHEET_PATTERN = re.compile(r"^(sheet\d+|工作表\d+)$", re.IGNORECASE)
# 表头保留行数
HEADER_ROWS = 3
# __tables__.xlsx 列顺序（与表头 ##var 行一致）
TABLE_COLUMNS = [
    "full_name",
    "value_type",
    "read_schema_from_file",
    "input",
    "index",
    "mode",
    "group",
    "comment",
    "tags",
    "output",
]


@dataclass
class SheetInfo:
    """解析出的 Luban 子表信息。"""

    file_rel: str
    sheet_name: str
    subtable_name: str
    index_field: str
    fields: list[str]


@dataclass
class TableRow:
    """__tables__.xlsx 一行数据。"""

    full_name: str = ""
    value_type: str = ""
    read_schema_from_file: str = ""
    input: str = ""
    index: str = ""
    mode: str = ""
    group: str = ""
    comment: str = ""
    tags: str = ""
    output: str = ""

    def to_list(self) -> list:
        return [
            self.full_name,
            self.value_type,
            self.read_schema_from_file,
            self.input,
            self.index,
            self.mode,
            self.group,
            self.comment,
            self.tags,
            self.output,
        ]

    @classmethod
    def from_list(cls, values: list) -> TableRow:
        padded = list(values) + [""] * len(TABLE_COLUMNS)
        return cls(**dict(zip(TABLE_COLUMNS, padded[: len(TABLE_COLUMNS)])))


@dataclass
class MergeResult:
    """合并结果统计。"""

    rows: list[TableRow] = field(default_factory=list)
    preserved: list[str] = field(default_factory=list)
    added: list[str] = field(default_factory=list)
    removed: list[str] = field(default_factory=list)
    migrated: list[str] = field(default_factory=list)
    renamed: list[str] = field(default_factory=list)


def normalize_input(value: str) -> str:
    """统一 input 路径格式，便于比较。"""
    return str(value or "").strip().replace("\\", "/")


def cell_to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_temp_file(path: Path) -> bool:
    return path.name.startswith("~$")


def should_skip_path(rel: Path) -> bool:
    """是否跳过该相对路径（目录在 SKIP_DIRS 中）。"""
    return any(part in SKIP_DIRS for part in rel.parts)


def is_luban_sheet(ws) -> bool:
    """判断 sheet 是否为 Luban 新格式数据表（首行 ##var，次行 ##type）。"""
    var_tag = cell_to_str(ws.cell(row=1, column=1).value)
    type_tag = cell_to_str(ws.cell(row=2, column=1).value)
    return var_tag == "##var" and type_tag == "##type"


def parse_var_fields(ws) -> tuple[list[str], str]:
    """从 ##var / ##type 行解析字段列表和 index 字段。"""
    var_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    fields = []
    for val in var_row[1:]:
        if val is None or str(val).strip() == "":
            continue
        fields.append(str(val).strip())
    index_field = fields[0] if fields else "id"
    return fields, index_field


def strip_hash_from_filename(filename: str) -> str:
    """去掉文件名开头的 #。"""
    stem = Path(filename).stem
    if stem.startswith("#"):
        stem = stem[1:]
    return stem


def filename_to_module_name(filename: str) -> str:
    """从文件名提取模块包名：#item.Item2.xlsx -> item；无 '.' 则返回空字符串。"""
    stem = strip_hash_from_filename(filename)
    if "." not in stem:
        return ""
    return to_snake_case(stem.split(".", 1)[0])


def filename_to_subtable_name(filename: str) -> str:
    """从 Excel 文件名推导子表名（取最后一个 '.' 后段）。"""
    stem = strip_hash_from_filename(filename)
    if "." in stem:
        return stem.rsplit(".", 1)[-1]
    return stem


def normalize_base_name(name: str) -> str:
    """xxx_base / base_xxx 归一化为 xxx（snake_case）。"""
    n = name.strip()
    lower = n.lower()
    if lower.endswith("_base"):
        return lower[: -len("_base")]
    if lower.startswith("base_"):
        return lower[len("base_") :]
    return lower


def to_snake_case(name: str) -> str:
    """转为 snake_case。"""
    if not name:
        return name
    if "_" in name and name == name.lower():
        return name.lower()
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", name)
    s = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1_\2", s)
    return s.replace("-", "_").lower()


def to_pascal_case(snake_name: str) -> str:
    """snake_case -> PascalCase。"""
    parts = [p for p in snake_name.split("_") if p]
    return "".join(p.capitalize() for p in parts)


def is_default_sheet_name(sheet_name: str) -> bool:
    return bool(DEFAULT_SHEET_PATTERN.match(sheet_name.strip()))


def resolve_subtable_name(sheet_name: str, file_name: str, single_default_sheet: bool) -> str:
    """确定子表名并做 base 归一化。"""
    if single_default_sheet and is_default_sheet_name(sheet_name):
        raw = filename_to_subtable_name(file_name)
    else:
        raw = sheet_name
    return normalize_base_name(to_snake_case(raw))


def resolve_module_namespace(file_rel: str) -> str:
    """
    从 Datas 相对路径提取模块命名空间（对应 Java 子包名）。
    优先级：
      1. 文件名含 '.'：#demo.demo.xlsx -> item
      2. 子目录名：item/foo.xlsx -> item
    """
    path = Path(file_rel)
    module = filename_to_module_name(path.name)
    if module:
        return module
    parts = path.parts
    if len(parts) > 1:
        return to_snake_case(parts[0])
    return ""


def parse_input_key(input_key: str) -> tuple[str, str, str]:
    """解析 input：sheet 名、文件相对路径、归一化子表名。"""
    normalized = normalize_input(input_key)
    if "@" not in normalized:
        return "", normalized, ""
    sheet, file_rel = normalized.split("@", 1)
    subtable = normalize_base_name(to_snake_case(filename_to_subtable_name(Path(file_rel).name)))
    return sheet, file_rel, subtable


def build_rename_map(
    existing_by_input: dict[str, TableRow],
    scanned_defaults: dict[str, TableRow],
) -> dict[str, str]:
    """
    识别同一逻辑表因文件重命名导致的 input 变化。
    返回 new_input -> old_input，例如 Sheet1@#demo.demo.xlsx <- Sheet1@#Item2.xlsx。
    """
    existing_keys = set(existing_by_input)
    scanned_keys = set(scanned_defaults)
    removed = existing_keys - scanned_keys
    added = scanned_keys - existing_keys

    added_index: dict[tuple[str, str], str] = {}
    for key in added:
        sheet, _, subtable = parse_input_key(key)
        if not sheet or not subtable:
            continue
        if (sheet, subtable) in added_index:
            continue
        added_index[(sheet, subtable)] = key

    rename_map: dict[str, str] = {}
    for old_key in removed:
        sheet, _, subtable = parse_input_key(old_key)
        new_key = added_index.get((sheet, subtable))
        if new_key:
            rename_map[new_key] = old_key
    return rename_map


def build_type_names(info: SheetInfo) -> tuple[str, str]:
    """生成带模块前缀的 full_name / value_type。"""
    pascal = to_pascal_case(info.subtable_name)
    table_name = f"{pascal}Table"
    bean_name = pascal
    module = resolve_module_namespace(info.file_rel)
    if module:
        return f"{module}.{table_name}", f"{module}.{bean_name}"
    return table_name, bean_name


def build_table_row(info: SheetInfo) -> TableRow:
    """为新增表生成默认行。"""
    full_name, value_type = build_type_names(info)
    return TableRow(
        full_name=full_name,
        value_type=value_type,
        read_schema_from_file="true",
        input=f"{info.sheet_name}@{info.file_rel.replace(chr(92), '/')}",
        index=info.index_field,
        mode="map",
        group="c,s",
        comment="",
        tags="",
        output=info.subtable_name,
    )


def should_migrate_namespace(existing: TableRow, default: TableRow) -> bool:
    """
    是否将扁平命名迁移为带模块前缀的命名。
    仅当现有行仍是扁平自动命名（未手动改过类名）时才迁移。
    """
    if "." in existing.full_name:
        return False
    if "." not in default.full_name:
        return False
    flat_table = default.full_name.rsplit(".", 1)[-1]
    flat_bean = default.value_type.rsplit(".", 1)[-1]
    return existing.full_name == flat_table and existing.value_type == flat_bean


def apply_namespace_migration(existing: TableRow, default: TableRow) -> TableRow:
    """保留手动维护字段，更新 full_name / value_type / input 为带模块前缀。"""
    return TableRow(
        full_name=default.full_name,
        value_type=default.value_type,
        read_schema_from_file=existing.read_schema_from_file or default.read_schema_from_file,
        input=default.input,
        index=existing.index or default.index,
        mode=existing.mode or default.mode,
        group=existing.group or default.group,
        comment=existing.comment,
        tags=existing.tags,
        output=existing.output or default.output,
    )


def scan_datas_dir(datas_dir: Path) -> list[SheetInfo]:
    results: list[SheetInfo] = []
    for path in sorted(datas_dir.rglob("*")):
        if not path.is_file():
            continue
        if path.suffix.lower() not in DATA_SUFFIXES:
            continue
        if path.name in SKIP_FILES or is_temp_file(path):
            continue

        rel_path = path.relative_to(datas_dir)
        if should_skip_path(rel_path):
            continue

        rel = rel_path.as_posix()
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"[WARN] 无法打开 {rel}: {e}")
            continue

        luban_sheets: list[tuple[str, list[str], str]] = []
        for name in wb.sheetnames:
            ws = wb[name]
            if is_luban_sheet(ws):
                fields, index_field = parse_var_fields(ws)
                luban_sheets.append((name, fields, index_field))

        wb.close()

        if not luban_sheets:
            print(f"[WARN] 跳过（无 Luban 数据 sheet）: {rel}")
            continue

        single_default = (
            len(luban_sheets) == 1 and is_default_sheet_name(luban_sheets[0][0])
        )

        for sheet_name, fields, index_field in luban_sheets:
            if not fields:
                print(f"[WARN] 跳过（无字段）: {rel} [{sheet_name}]")
                continue
            subtable = resolve_subtable_name(sheet_name, path.name, single_default)
            results.append(
                SheetInfo(
                    file_rel=rel,
                    sheet_name=sheet_name,
                    subtable_name=subtable,
                    index_field=index_field,
                    fields=fields,
                )
            )
    return results


def deduplicate_rows(sheets: list[SheetInfo]) -> tuple[list[SheetInfo], list[str]]:
    """按 input 去重，保留首次出现。"""
    seen: dict[str, SheetInfo] = {}
    warnings: list[str] = []
    for info in sheets:
        input_key = normalize_input(build_table_row(info).input)
        if input_key in seen:
            prev = seen[input_key]
            warnings.append(
                f"input 重复 '{input_key}': "
                f"{prev.file_rel}[{prev.sheet_name}] 与 {info.file_rel}[{info.sheet_name}]"
            )
            continue
        seen[input_key] = info
    return list(seen.values()), warnings


def load_existing_rows(output_path: Path) -> dict[str, TableRow]:
    """读取已有 __tables__.xlsx 数据行，按 input 索引。"""
    if not output_path.exists():
        return {}

    wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    ws = wb.active
    existing: dict[str, TableRow] = {}

    for row_idx in range(HEADER_ROWS + 1, ws.max_row + 1):
        values = [cell_to_str(ws.cell(row=row_idx, column=col).value) for col in range(2, 2 + len(TABLE_COLUMNS))]
        if not any(values):
            continue
        row = TableRow.from_list(values)
        key = normalize_input(row.input)
        if not key:
            continue
        if key in existing:
            print(f"[WARN] __tables__.xlsx 存在重复 input，保留首次出现: {key}")
            continue
        existing[key] = row

    wb.close()
    return existing


def merge_table_rows(
    scanned_sheets: list[SheetInfo],
    existing_by_input: dict[str, TableRow],
) -> MergeResult:
    """
    增量合并：
      - input 已存在 -> 保留原有行（含手动修改）
      - input 新增     -> 使用默认规则生成
      - 文件重命名但子表不变 -> 保留原行并更新 input / 命名空间
      - 原有 input 不在扫描结果中 -> 删除
    """
    result = MergeResult()
    scanned_defaults: dict[str, TableRow] = {}

    for info in scanned_sheets:
        default_row = build_table_row(info)
        key = normalize_input(default_row.input)
        scanned_defaults[key] = default_row

    scanned_inputs = set(scanned_defaults)
    rename_map = build_rename_map(existing_by_input, scanned_defaults)
    consumed_old_keys: set[str] = set()

    for key, default_row in scanned_defaults.items():
        if key in existing_by_input:
            preserved = existing_by_input[key]
            if should_migrate_namespace(preserved, default_row):
                preserved = apply_namespace_migration(preserved, default_row)
                result.migrated.append(key)
            result.rows.append(preserved)
            result.preserved.append(key)
        elif key in rename_map:
            old_key = rename_map[key]
            consumed_old_keys.add(old_key)
            preserved = existing_by_input[old_key]
            if should_migrate_namespace(preserved, default_row):
                preserved = apply_namespace_migration(preserved, default_row)
                result.migrated.append(key)
            else:
                preserved = TableRow(
                    full_name=preserved.full_name,
                    value_type=preserved.value_type,
                    read_schema_from_file=preserved.read_schema_from_file or default_row.read_schema_from_file,
                    input=default_row.input,
                    index=preserved.index or default_row.index,
                    mode=preserved.mode or default_row.mode,
                    group=preserved.group or default_row.group,
                    comment=preserved.comment,
                    tags=preserved.tags,
                    output=preserved.output or default_row.output,
                )
            result.rows.append(preserved)
            result.preserved.append(key)
            result.renamed.append(f"{old_key} -> {key}")
        else:
            result.rows.append(default_row)
            result.added.append(key)

    for key in existing_by_input:
        if key not in scanned_inputs and key not in consumed_old_keys:
            result.removed.append(key)

    return result


def write_tables_xlsx(output_path: Path, rows: list[TableRow], dry_run: bool) -> None:
    if dry_run:
        return

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        if ws.max_row > HEADER_ROWS:
            ws.delete_rows(HEADER_ROWS + 1, ws.max_row - HEADER_ROWS)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["##var"] + TABLE_COLUMNS)
        ws.append(["##"] + [""] * len(TABLE_COLUMNS))
        ws.append(["##"] + [""] * len(TABLE_COLUMNS))

    for row in rows:
        ws.append([""] + row.to_list())

    wb.save(output_path)


def print_report(
    merge: MergeResult,
    scanned_sheets: list[SheetInfo],
    warnings: list[str],
) -> None:
    print("=" * 72)
    print("  Luban __tables__.xlsx 增量更新报告")
    print("=" * 72)

    if warnings:
        print("\n[WARN] 冲突/重复：")
        for w in warnings:
            print(f"  - {w}")

    print(f"\n扫描子表: {len(scanned_sheets)}  |  写入行数: {len(merge.rows)}")
    print(
        f"  保留: {len(merge.preserved)}  |  新增: {len(merge.added)}  |  "
        f"删除: {len(merge.removed)}  |  重命名: {len(merge.renamed)}  |  "
        f"命名空间迁移: {len(merge.migrated)}"
    )

    if merge.renamed:
        print("\n[重命名] input 已更新，保留原表定义")
        for item in merge.renamed:
            print(f"  ~ {item}")

    if merge.migrated:
        print("\n[命名空间迁移] full_name / value_type 已补充模块前缀")
        for key in merge.migrated:
            print(f"  ~ {key}")

    if merge.added:
        print("\n[新增]")
        for key in merge.added:
            print(f"  + {key}")

    if merge.removed:
        print("\n[删除]")
        for key in merge.removed:
            print(f"  - {key}")

    print(f"\n{'状态':<8} {'full_name':<28} {'value_type':<20} {'output':<16} input")
    print("-" * 120)
    preserved_set = set(merge.preserved)
    migrated_set = set(merge.migrated)
    for row in merge.rows:
        key = normalize_input(row.input)
        if key in migrated_set:
            status = "迁移"
        elif key in preserved_set:
            status = "保留"
        else:
            status = "新增"
        print(
            f"{status:<8} {row.full_name:<28} {row.value_type:<20} "
            f"{row.output:<16} {row.input}"
        )
    print()


def main() -> int:
    parser = argparse.ArgumentParser(description="扫描 Datas 并增量更新 __tables__.xlsx")
    parser.add_argument(
        "--datas",
        type=Path,
        default=Path(__file__).resolve().parent / "Datas",
        help="Luban 数据目录（默认：脚本同级 Datas/）",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出 __tables__.xlsx 路径（默认：{datas}/__tables__.xlsx）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="仅打印结果，不写入文件",
    )
    args = parser.parse_args()

    datas_dir = args.datas.resolve()
    output_path = (args.output or datas_dir / "__tables__.xlsx").resolve()

    if not datas_dir.is_dir():
        print(f"[ERROR] 数据目录不存在: {datas_dir}")
        return 1

    print(f"扫描目录: {datas_dir}")
    all_sheets = scan_datas_dir(datas_dir)
    unique_sheets, warnings = deduplicate_rows(all_sheets)

    existing_by_input = load_existing_rows(output_path)
    merge = merge_table_rows(unique_sheets, existing_by_input)

    print_report(merge, unique_sheets, warnings)

    if args.dry_run:
        print("[DRY-RUN] 未写入文件")
        return 0

    try:
        write_tables_xlsx(output_path, merge.rows, dry_run=False)
        print(f"[OK] 已写入: {output_path}")
    except PermissionError:
        print(f"[ERROR] 无法写入 {output_path}，请关闭 Excel 后重试")
        return 1
    except Exception as e:
        print(f"[ERROR] 写入失败: {e}")
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
